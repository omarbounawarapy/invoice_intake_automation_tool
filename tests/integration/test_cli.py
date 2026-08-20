"""Integration tests for the invoice-tool CLI.

Calls invoice_intake_automation_tool.cli.main() directly with an argv list
and captures stdout/stderr via capsys, rather than shelling out to a
subprocess -- exercises the exact same code path a subprocess would, at a
fraction of the cost.
"""

from __future__ import annotations

import json

import pytest

from invoice_intake_automation_tool.cli import main


class TestExtractCommand:
    def test_default_format_is_readable_text(self, example_invoices_dir, capsys):
        exit_code = main(["extract", str(example_invoices_dir / "INV-2026-0001.pdf")])
        out = capsys.readouterr().out
        assert exit_code == 0
        assert "INV-2026-0001" in out
        assert "{" not in out  # not JSON

    def test_json_format(self, example_invoices_dir, capsys):
        exit_code = main(["extract", str(example_invoices_dir / "INV-2026-0001.pdf"), "--format", "json"])
        out = capsys.readouterr().out
        assert exit_code == 0
        payload = json.loads(out)
        assert payload["invoice_id"] == "INV-2026-0001"

    def test_csv_format(self, example_invoices_dir, capsys):
        exit_code = main(["extract", str(example_invoices_dir / "INV-2026-0001.pdf"), "--format", "csv"])
        out = capsys.readouterr().out
        assert exit_code == 0
        assert out.splitlines()[0].startswith("invoice_id,")

    def test_xlsx_format_written_to_file(self, example_invoices_dir, tmp_path):
        from openpyxl import load_workbook

        output_path = tmp_path / "result.xlsx"
        exit_code = main(
            ["extract", str(example_invoices_dir / "INV-2026-0001.pdf"), "--format", "xlsx", "--output", str(output_path)]
        )
        assert exit_code == 0
        wb = load_workbook(output_path)
        assert wb.sheetnames == ["Invoice Summary", "Line Items"]

    def test_xlsx_format_refuses_an_interactive_terminal(self, example_invoices_dir, monkeypatch, capsys):
        monkeypatch.setattr("sys.stdout.isatty", lambda: True)
        exit_code = main(["extract", str(example_invoices_dir / "INV-2026-0001.pdf"), "--format", "xlsx"])
        captured = capsys.readouterr()
        assert exit_code == 1
        assert "--output" in captured.err

    def test_output_flag_writes_a_file_instead_of_stdout(self, example_invoices_dir, tmp_path, capsys):
        output_path = tmp_path / "result.json"
        exit_code = main(
            [
                "extract",
                str(example_invoices_dir / "INV-2026-0001.pdf"),
                "--format",
                "json",
                "--output",
                str(output_path),
            ]
        )
        assert exit_code == 0
        assert output_path.exists()
        payload = json.loads(output_path.read_text())
        assert payload["invoice_id"] == "INV-2026-0001"
        assert capsys.readouterr().out == ""  # nothing written to stdout

    def test_missing_file_exits_nonzero_with_a_clean_message(self, tmp_path, capsys):
        exit_code = main(["extract", str(tmp_path / "missing.pdf")])
        captured = capsys.readouterr()
        assert exit_code == 1
        assert "Error (input)" in captured.err
        assert "Traceback" not in captured.err

    def test_strict_flag_fails_on_a_known_total_mismatch(self, example_invoices_dir, capsys):
        exit_code = main(["extract", str(example_invoices_dir / "INV-2026-0002.pdf"), "--strict"])
        assert exit_code == 1

    def test_strict_flag_does_not_affect_a_correct_invoice(self, example_invoices_dir, capsys):
        exit_code = main(["extract", str(example_invoices_dir / "INV-2026-0001.pdf"), "--strict"])
        assert exit_code == 0

    def test_without_strict_a_mismatch_still_exits_zero(self, example_invoices_dir, capsys):
        # A reconciliation mismatch is diagnostic, not a processing
        # failure -- only --strict should turn it into a non-zero exit.
        exit_code = main(["extract", str(example_invoices_dir / "INV-2026-0002.pdf")])
        assert exit_code == 0


class TestBatchCommand:
    def test_default_output_is_a_per_file_summary(self, example_invoices_dir, capsys):
        exit_code = main(["batch", str(example_invoices_dir)])
        out = capsys.readouterr().out
        assert exit_code == 0
        assert "INV-2026-0001.pdf" in out
        assert "5 processed, 0 failed" in out

    def test_json_format_combines_every_invoice(self, example_invoices_dir, capsys):
        exit_code = main(["batch", str(example_invoices_dir), "--format", "json"])
        out = capsys.readouterr().out
        assert exit_code == 0
        payload = json.loads(out)
        assert len(payload) == 5

    def test_xlsx_format_combines_every_invoice(self, example_invoices_dir, tmp_path):
        from openpyxl import load_workbook

        output_path = tmp_path / "batch.xlsx"
        exit_code = main(["batch", str(example_invoices_dir), "--format", "xlsx", "--output", str(output_path)])
        assert exit_code == 0
        wb = load_workbook(output_path)
        assert wb["Invoice Summary"].max_row == 6  # header + 5 invoices

    def test_nonexistent_directory_exits_nonzero(self, tmp_path, capsys):
        exit_code = main(["batch", str(tmp_path / "missing-dir")])
        assert exit_code == 1

    def test_strict_flag_fails_when_any_invoice_is_flagged(self, example_invoices_dir):
        exit_code = main(["batch", str(example_invoices_dir), "--strict"])
        assert exit_code == 1  # two of the five example invoices have a total mismatch

    def test_pattern_option_is_respected(self, example_invoices_dir, capsys):
        exit_code = main(["batch", str(example_invoices_dir), "--pattern", "INV-2026-0001.pdf"])
        out = capsys.readouterr().out
        assert exit_code == 0
        assert "1 processed, 0 failed" in out


class TestArgumentErrors:
    def test_missing_subcommand_exits_nonzero(self, capsys):
        with pytest.raises(SystemExit) as exc_info:
            main([])
        assert exc_info.value.code != 0

    def test_unknown_format_choice_exits_nonzero(self, example_invoices_dir):
        with pytest.raises(SystemExit) as exc_info:
            main(["extract", str(example_invoices_dir / "INV-2026-0001.pdf"), "--format", "xml"])
        assert exc_info.value.code != 0
