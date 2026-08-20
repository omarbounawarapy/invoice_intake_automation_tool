"""Unit tests for invoice_intake_automation_tool.serialize.

Checks structure/round-tripping, not exact byte-for-byte formatting (the
CLI regression tests separately pin the real example invoices' text
output).
"""

from __future__ import annotations

import csv
import io
import json
from decimal import Decimal

from mined_invoices import base_mined_invoice
from openpyxl import load_workbook

from invoice_intake_automation_tool.serialize import CSV_COLUMNS, XLSX_SUMMARY_COLUMNS, to_csv, to_json, to_text, to_xlsx
from invoice_intake_automation_tool.transform import transform_invoice


def _invoice(**overrides):
    return transform_invoice(base_mined_invoice(**overrides))


class TestToJson:
    def test_single_invoice_round_trips_through_json(self):
        invoice = _invoice()
        payload = json.loads(to_json(invoice))
        assert payload["invoice_id"] == "INV-TEST-0001"
        assert payload["total"] == "1200.00"  # Decimal serialized as an exact string, not a float

    def test_list_of_invoices_serializes_as_a_json_array(self):
        invoices = [_invoice(), _invoice(invoice_id="INV-TEST-0002")]
        payload = json.loads(to_json(invoices))
        assert isinstance(payload, list)
        assert [item["invoice_id"] for item in payload] == ["INV-TEST-0001", "INV-TEST-0002"]

    def test_non_ascii_names_are_not_escaped(self):
        invoice = _invoice(vendor="Café Müller GmbH")
        assert "Café Müller GmbH" in to_json(invoice)


class TestToCsv:
    def test_header_matches_declared_columns(self):
        rows = list(csv.reader(io.StringIO(to_csv(_invoice()))))
        assert rows[0] == CSV_COLUMNS

    def test_one_row_per_line_item(self):
        invoice = _invoice(
            line_items=[
                {"description": "A", "quantity": "1", "unit_price": "10.00", "amount": "10.00", "vat_rate": None},
                {"description": "B", "quantity": "1", "unit_price": "5.00", "amount": "5.00", "vat_rate": None},
            ]
        )
        rows = list(csv.DictReader(io.StringIO(to_csv(invoice))))
        assert len(rows) == 2
        assert [r["line_description"] for r in rows] == ["A", "B"]
        # Invoice-level fields repeat identically on every line-item row.
        assert {r["invoice_id"] for r in rows} == {"INV-TEST-0001"}

    def test_batch_concatenates_multiple_invoices(self):
        invoices = [_invoice(), _invoice(invoice_id="INV-TEST-0002")]
        rows = list(csv.DictReader(io.StringIO(to_csv(invoices))))
        assert {r["invoice_id"] for r in rows} == {"INV-TEST-0001", "INV-TEST-0002"}

    def test_money_values_have_no_thousands_separators(self):
        # CSV values must be directly usable in spreadsheet formulas.
        invoice = _invoice()
        rows = list(csv.DictReader(io.StringIO(to_csv(invoice))))
        assert rows[0]["subtotal"] == "1000.00"
        assert "," not in rows[0]["subtotal"]


class TestToXlsx:
    def _load(self, data: bytes):
        return load_workbook(io.BytesIO(data))

    def test_returns_bytes(self):
        assert isinstance(to_xlsx(_invoice()), bytes)

    def test_has_summary_and_line_items_sheets(self):
        wb = self._load(to_xlsx(_invoice()))
        assert wb.sheetnames == ["Invoice Summary", "Line Items"]

    def test_summary_sheet_header_matches_declared_columns(self):
        wb = self._load(to_xlsx(_invoice()))
        header = [cell.value for cell in next(wb["Invoice Summary"].iter_rows(min_row=1, max_row=1))]
        assert header == XLSX_SUMMARY_COLUMNS

    def test_summary_sheet_has_one_row_per_invoice(self):
        invoices = [_invoice(), _invoice(invoice_id="INV-TEST-0002")]
        wb = self._load(to_xlsx(invoices))
        ws = wb["Invoice Summary"]
        assert ws.max_row == 3  # header + 2 invoices
        ids = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert ids == ["INV-TEST-0001", "INV-TEST-0002"]

    def test_line_items_sheet_matches_csv_columns(self):
        wb = self._load(to_xlsx(_invoice()))
        header = [cell.value for cell in next(wb["Line Items"].iter_rows(min_row=1, max_row=1))]
        assert header == CSV_COLUMNS

    def test_monetary_values_are_numeric_cells(self):
        wb = self._load(to_xlsx(_invoice()))
        ws = wb["Invoice Summary"]
        subtotal_cell = next(ws.iter_rows(min_row=2, max_row=2))[XLSX_SUMMARY_COLUMNS.index("subtotal")]
        assert isinstance(subtotal_cell.value, (int, float))
        assert subtotal_cell.value == 1000.00

    def test_single_invoice_and_list_produce_the_same_shape(self):
        # to_xlsx(invoice) should behave like to_xlsx([invoice]).
        single = self._load(to_xlsx(_invoice()))
        batch = self._load(to_xlsx([_invoice()]))
        assert single["Invoice Summary"].max_row == batch["Invoice Summary"].max_row


class TestToText:
    def test_contains_key_figures(self):
        text = to_text(_invoice())
        assert "INV-TEST-0001" in text
        assert "1,000.00" in text  # subtotal, human-formatted with a thousands separator
        assert "1,200.00" in text  # total

    def test_flags_consistency_issues(self):
        text = to_text(_invoice(rendered_total="1300.00"))
        assert "TOTAL ERROR" in text
        assert "differs from computed total" in text

    def test_correct_invoice_has_no_warning_line(self):
        text = to_text(_invoice())
        assert "!" not in text

    def test_long_consistency_note_is_wrapped(self):
        # Both subtotal and total mismatching produces a two-sentence note
        # that is long enough to need wrapping to stay readable.
        text = to_text(_invoice(rendered_subtotal="1500.00", rendered_total="2000.00"))
        assert all(len(line) < 80 for line in text.splitlines())

    def test_long_description_is_truncated_not_wrapped(self):
        text = to_text(
            _invoice(
                line_items=[
                    {
                        "description": "A" * 80,
                        "quantity": "1",
                        "unit_price": "1000.00",
                        "amount": "1000.00",
                        "vat_rate": None,
                    }
                ]
            )
        )
        # No single line should run away with the full 80-character
        # description -- it must be clipped to keep the summary readable.
        assert all(len(line) < 80 for line in text.splitlines())
