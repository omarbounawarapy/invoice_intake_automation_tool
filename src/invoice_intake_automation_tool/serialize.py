"""Serialize validated :class:`Invoice` objects to the project's supported
output formats: JSON, CSV, XLSX, and a human-readable terminal summary.

CSV/XLSX flattening rule (see docs/data-model.md for the full rationale):
one row per **line item**, with invoice-level fields repeated on every
row. This is the conventional shape for invoice data destined for a
spreadsheet -- it can be filtered, pivoted, and summed directly without a
second parsing step, unlike a CSV that embeds a nested list/dict per cell.
XLSX additionally gets an "Invoice Summary" sheet (one row per invoice)
alongside the same "Line Items" sheet, since a spreadsheet reviewer
generally wants the invoice-level totals reviewable on their own.
"""

from __future__ import annotations

import csv
import io
import textwrap
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from .models import DiscountType, Invoice

#: Column order for CSV output. Kept as a single explicit list (rather than
#: derived from the model) so the on-disk contract stays stable even if
#: unrelated fields are added to Invoice later.
CSV_COLUMNS = [
    "invoice_id",
    "invoice_date",
    "due_date",
    "vendor",
    "vendor_country",
    "recipient",
    "recipient_country",
    "currency",
    "line_number",
    "line_description",
    "line_quantity",
    "line_unit_price",
    "line_amount",
    "line_vat_rate",
    "subtotal",
    "discount_type",
    "discount_value",
    "discount_amount",
    "vat_rate",
    "vat_amount",
    "total",
    "payment_terms",
    "consistency",
    "consistency_note",
    "source_file",
]


def to_json(invoices: Invoice | list[Invoice], *, indent: int | None = 2) -> str:
    """Serialize one invoice, or a list of invoices, to JSON.

    Decimal fields are emitted as JSON strings (e.g. ``"210215.90"``), not
    floats, so exact monetary values survive round-tripping.
    """
    if isinstance(invoices, list):
        # Pydantic does not expose a list-of-models dump helper directly;
        # dumping each model to its JSON-mode dict keeps Decimal/date/Enum
        # encoding identical to the single-invoice path.
        import json

        payload = [invoice.model_dump(mode="json") for invoice in invoices]
        return json.dumps(payload, indent=indent, ensure_ascii=False)
    return invoices.model_dump_json(indent=indent)


def _line_rows(invoice: Invoice) -> list[dict[str, object]]:
    rows = []
    for number, item in enumerate(invoice.line_items, start=1):
        rows.append(
            {
                "invoice_id": invoice.invoice_id,
                "invoice_date": invoice.invoice_date.isoformat(),
                "due_date": invoice.due_date.isoformat() if invoice.due_date else "",
                "vendor": invoice.vendor,
                "vendor_country": invoice.vendor_country or "",
                "recipient": invoice.recipient,
                "recipient_country": invoice.recipient_country or "",
                "currency": invoice.currency or "",
                "line_number": number,
                "line_description": item.description,
                "line_quantity": item.quantity,
                "line_unit_price": item.unit_price,
                "line_amount": item.amount,
                "line_vat_rate": item.vat_rate if item.vat_rate is not None else "",
                "subtotal": invoice.subtotal,
                "discount_type": invoice.discount.type.value if invoice.discount else "",
                "discount_value": invoice.discount.value if invoice.discount else "",
                "discount_amount": invoice.discount_amount,
                "vat_rate": invoice.vat_rate if invoice.vat_rate is not None else "",
                "vat_amount": invoice.vat_amount,
                "total": invoice.total,
                "payment_terms": invoice.payment_terms or "",
                "consistency": invoice.consistency.value,
                "consistency_note": invoice.consistency_note or "",
                "source_file": invoice.source_file or "",
            }
        )
    return rows


def to_csv(invoices: Invoice | list[Invoice]) -> str:
    """Flatten one invoice, or a batch of invoices, to CSV (one row per
    line item; see module docstring)."""
    if not isinstance(invoices, list):
        invoices = [invoices]

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=CSV_COLUMNS, lineterminator="\n")
    writer.writeheader()
    for invoice in invoices:
        writer.writerows(_line_rows(invoice))
    return buffer.getvalue()


#: Columns for the XLSX "Invoice Summary" sheet -- see to_xlsx.
XLSX_SUMMARY_COLUMNS = [
    "invoice_id",
    "invoice_date",
    "due_date",
    "vendor",
    "vendor_country",
    "recipient",
    "recipient_country",
    "currency",
    "subtotal",
    "discount_amount",
    "vat_rate",
    "vat_amount",
    "total",
    "consistency",
    "consistency_note",
    "payment_terms",
    "source_file",
]


def _summary_row(invoice: Invoice) -> list[object]:
    return [
        invoice.invoice_id,
        invoice.invoice_date.isoformat(),
        invoice.due_date.isoformat() if invoice.due_date else None,
        invoice.vendor,
        invoice.vendor_country,
        invoice.recipient,
        invoice.recipient_country,
        invoice.currency,
        invoice.subtotal,
        invoice.discount_amount,
        invoice.vat_rate,
        invoice.vat_amount,
        invoice.total,
        invoice.consistency.value,
        invoice.consistency_note,
        invoice.payment_terms,
        invoice.source_file,
    ]


def _write_sheet(ws: Worksheet, columns: list[str], rows: list[list[object]]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        lengths = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
        width = min(max(lengths, default=10) + 2, 40)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def to_xlsx(invoices: Invoice | list[Invoice]) -> bytes:
    """Build a review-friendly Excel workbook: one row per invoice on an
    "Invoice Summary" sheet, and the same line-item flattening as
    :func:`to_csv` on a "Line Items" sheet.

    Monetary values are written as native numeric cells (via ``Decimal``,
    which ``openpyxl`` stores as a plain number) so they are directly
    usable in spreadsheet formulas -- unlike the JSON output, which keeps
    exact Decimal *strings*, this format takes on the ordinary
    floating-point representation Excel itself uses for every number.
    """
    if not isinstance(invoices, list):
        invoices = [invoices]

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Invoice Summary"
    _write_sheet(summary_ws, XLSX_SUMMARY_COLUMNS, [_summary_row(inv) for inv in invoices])

    items_ws = workbook.create_sheet("Line Items")
    item_rows = [[row[column] for column in CSV_COLUMNS] for inv in invoices for row in _line_rows(inv)]
    _write_sheet(items_ws, CSV_COLUMNS, item_rows)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_WIDTH = 78
_DESC_WIDTH = 38


def _fmt_money(value: Decimal, currency: str | None) -> str:
    formatted = f"{value:,.2f}"
    return f"{formatted} {currency}" if currency else formatted


def _row(label: str, amount: str, *, width: int = _WIDTH) -> str:
    """Right-align ``amount`` at ``width``, left-padding with a space
    after ``label`` -- never truncates, so an unusually long label simply
    pushes the amount further right rather than corrupting the layout."""
    gap = max(1, width - len(label) - len(amount))
    return f"{label}{' ' * gap}{amount}"


def _discount_label(discount) -> str:
    if discount.type is DiscountType.AMOUNT:
        label = "Discount"
    else:
        label = f"Discount ({discount.value:.0%})"
    if discount.conditional:
        label += " [offered, not applied]"
    return label + ":"


def to_text(invoice: Invoice) -> str:
    """A readable terminal summary -- the default ``extract`` output
    when no ``--format`` is given."""
    lines: list[str] = []

    status = invoice.consistency.value.replace("_", " ").upper()
    lines.append(_row(f"Invoice {invoice.invoice_id}", status))
    lines.append(
        f"{invoice.vendor} ({invoice.vendor_country or '?'})  ->  "
        f"{invoice.recipient} ({invoice.recipient_country or '?'})"
    )

    date_line = f"Date: {invoice.invoice_date}"
    if invoice.due_date:
        date_line += f"    Due: {invoice.due_date}"
    if invoice.currency:
        date_line += f"    Currency: {invoice.currency}"
    lines.append(date_line)
    if invoice.is_credit_note:
        lines.append("(credit note)")
    lines.append("-" * _WIDTH)

    lines.append("Line items")
    for item in invoice.line_items:
        desc = item.description
        if len(desc) > _DESC_WIDTH:
            desc = desc[: _DESC_WIDTH - 3] + "..."
        detail = f"{item.quantity} x {item.unit_price:,.2f} = {item.amount:,.2f}"
        lines.append(_row(f"  {desc}", detail))

    lines.append("-" * _WIDTH)

    lines.append(_row("Subtotal:", _fmt_money(invoice.subtotal, invoice.currency)))
    if invoice.discount:
        lines.append(
            _row(_discount_label(invoice.discount), _fmt_money(-invoice.discount_amount, invoice.currency))
        )
    rate_label = f"VAT ({invoice.vat_rate:.0%}):" if invoice.vat_rate is not None else "VAT:"
    lines.append(_row(rate_label, _fmt_money(invoice.vat_amount, invoice.currency)))
    lines.append("=" * _WIDTH)
    lines.append(_row("Total:", _fmt_money(invoice.total, invoice.currency)))

    if invoice.consistency_note:
        lines.append("")
        lines.append(textwrap.fill(f"! {invoice.consistency_note}", width=_WIDTH, subsequent_indent="  "))

    return "\n".join(lines)
